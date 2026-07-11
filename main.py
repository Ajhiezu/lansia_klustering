"""
main.py - Entry point project clustering lansia
Ganti INPUT_FILE di config.py untuk data bulan berikutnya
"""

import shutil
import sys
import logging
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Setup path ────────────────────────────────────────────────────────────
PROJECT_DIR = Path(__file__).parent
sys.path.insert(0, str(PROJECT_DIR))

from app.core.config import (
    INPUT_FILE, OUTPUT_DIR, VIS_DIR, LOG_DIR, CSV_DIR,
    N_CLUSTERS, CLUSTER_FEATURES
)
from app.core.utils import setup_logger
from app.core.preprocessing import run_preprocessing
from app.core.clustering import run_clustering
from visualization import run_all_visualizations


# ══════════════════════════════════════════════════════════════════════════════
# Helpers Excel styling
# ══════════════════════════════════════════════════════════════════════════════

def _header_style(ws, row: int, fill_hex: str = "1F4E79"):
    """Beri warna header baris ke-row pada worksheet."""
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[row]:
        if cell.value is not None or True:
            cell.fill   = fill
            cell.font   = font
            cell.alignment = align
            cell.border = border


def _autowidth(ws, max_width: int = 40):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, max_width
        )


def _df_to_sheet(ws, df: pd.DataFrame, fill_hex: str = "1F4E79"):
    """Tulis DataFrame ke worksheet dengan header bergaya."""
    ws.append(list(df.columns))
    _header_style(ws, 1, fill_hex)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    _autowidth(ws)


# ══════════════════════════════════════════════════════════════════════════════
# Output 1 – hasil_preprocessing.xlsx
# ══════════════════════════════════════════════════════════════════════════════

def save_preprocessing_excel(df: pd.DataFrame, path: Path):
    """Simpan hasil preprocessing ke Excel multi-sheet."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet GABUNGAN
    ws_all = wb.create_sheet("GABUNGAN")
    _df_to_sheet(ws_all, df)

    # Sheet per desa
    for desa, grp in df.groupby("desa"):
        ws = wb.create_sheet(str(desa)[:31])  # max 31 karakter
        _df_to_sheet(ws, grp.reset_index(drop=True), fill_hex="2E7D32")

    wb.save(path)
    logger.info(f"Disimpan: {path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# Output 2 – hasil_clustering.xlsx
# ══════════════════════════════════════════════════════════════════════════════

def save_clustering_excel(df: pd.DataFrame, df_centroids: pd.DataFrame,
                          path: Path):
    """Simpan hasil clustering dengan centroid ke Excel."""
    wb = Workbook()
    wb.remove(wb.active)

    # HASIL_CLUSTER – semua data + kolom cluster
    ws_main = wb.create_sheet("HASIL_CLUSTER")
    _df_to_sheet(ws_main, df, fill_hex="1A237E")

    # Per desa
    for desa, grp in df.groupby("desa"):
        ws = wb.create_sheet(str(desa)[:31])
        _df_to_sheet(ws, grp.reset_index(drop=True), fill_hex="4A148C")

    # GABUNGAN
    ws_gab = wb.create_sheet("GABUNGAN")
    _df_to_sheet(ws_gab, df, fill_hex="1B5E20")

    # CENTROID_CLUSTER
    ws_centroid = wb.create_sheet("CENTROID_CLUSTER")
    df_c = df_centroids.reset_index().rename(columns={"index": "cluster"})
    _df_to_sheet(ws_centroid, df_c, fill_hex="BF360C")

    wb.save(path)
    logger.info(f"Disimpan: {path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# Output 3 – hasil_preprocessing_scaled.xlsx
# ══════════════════════════════════════════════════════════════════════════════

def save_scaled_excel(df_scaled: pd.DataFrame, cluster_labels: pd.Series,
                      path: Path):
    df_s = df_scaled.copy()
    df_s["cluster"] = cluster_labels.values
    wb = Workbook()
    ws = wb.active
    ws.title = "SCALED_FEATURES"
    _df_to_sheet(ws, df_s, fill_hex="006064")
    wb.save(path)
    logger.info(f"Disimpan: {path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# Output 4 – evaluasi_clustering.txt
# ══════════════════════════════════════════════════════════════════════════════

def save_evaluation_txt(metrics: dict, interpretasi: str, path: Path):
    """Tulis laporan evaluasi ke file teks."""
    sil  = metrics["silhouette_score"]
    dbi  = metrics["davies_bouldin_index"]
    dist = metrics["jumlah_data_per_cluster"]
    ch = metrics["calinski_harabasz_score"]

    # Interpretasi kualitas
    if sil >= 0.5:
        kualitas_sil = "BAIK – cluster terpisah dengan jelas"
    elif sil >= 0.25:
        kualitas_sil = "CUKUP – cluster agak tumpang tindih"
    else:
        kualitas_sil = "KURANG – cluster saling tumpang tindih"

    kualitas_dbi = "BAIK – cluster kompak & terpisah" if dbi <= 1.0 else \
                   "CUKUP" if dbi <= 2.0 else "KURANG"

    lines = [
        "=" * 60,
        "EVALUASI CLUSTERING K-MEANS – DATA LANSIA PUSKESMAS MAESAN",
        "=" * 60,
        "",
        f"Jumlah Cluster         : {N_CLUSTERS}",
        f"Silhouette Score       : {sil:.4f}  → {kualitas_sil}",
        f"Davies-Bouldin Index   : {dbi:.4f}  → {kualitas_dbi}",
        f"Inertia (WCSS)         : {metrics['inertia']:.2f}",
        f"Calinski-Harabasz      : {ch:.4f}",
        "",
        "Jumlah Data per Cluster:",
    ]
    for cl, cnt in sorted(dist.items()):
        lines.append(f"  Cluster {cl} : {cnt} lansia")

    lines.append("")
    lines.append("CATATAN INTERPRETASI:")
    lines.append(
        "  Silhouette Score mendekati 1.0 = clustering sangat baik."
    )
    lines.append(
        "  Davies-Bouldin mendekati 0.0  = cluster kompak & berjauhan."
    )
    lines.append(interpretasi)

    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Disimpan: {path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# Output 5 – validation_summary.txt
# ══════════════════════════════════════════════════════════════════════════════

def save_validation_txt(stats: dict, path: Path):
    lines = [
        "=" * 60,
        "VALIDATION SUMMARY – PREPROCESSING DATA LANSIA",
        "=" * 60,
        "",
        f"Sheet diproses         : {stats.get('sheets_diproses', '-')}",
        f"Total baris mentah     : {stats.get('total_baris_mentah', '-')}",
        f"Total data valid       : {stats.get('total_valid', '-')}",
        f"Baris dihapus          : {stats.get('baris_dihapus', '-')}",
        "",
        f"Umur invalid (→ NaN)   : {stats.get('umur_invalid', '-')}",
        f"IMT invalid  (→ NaN)   : {stats.get('imt_invalid', '-')}",
        f"TD invalid   (→ NaN)   : {stats.get('td_invalid', '-')}",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Disimpan: {path.name}")


# ══════════════════════════════════════════════════════════════════════════════
# Output 6 – GIS CSV
# ══════════════════════════════════════════════════════════════════════════════

def save_gis_csv(df: pd.DataFrame, path: Path):
    """
    Export CSV siap QGIS.
    """
    cols_gis = ["nama", "nik", "kecamatan", "desa", "umur", "jenis_kelamin",
                "cluster", "cluster_label", "imt", "sistolik", "diastolik",
                "kolesterol", "gds_1", "gds_2", "gdp", "gd2pp",
                "riwayat_hipertensi", "gangguan_kognitif", "malnutrisi", "depresi"]
    cols_gis = [c for c in cols_gis if c in df.columns]
    df[cols_gis].to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"Disimpan: {path.name} (siap QGIS)")

def save_k_evaluation_excel(df_eval: pd.DataFrame, path: Path):

    wb = Workbook()
    ws = wb.active
    ws.title = "EVALUASI_K"

    _df_to_sheet(ws, df_eval, fill_hex="4E342E")

    wb.save(path)

    logger.info(f"Disimpan: {path.name}")       


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY AKHIR
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(stats: dict, metrics: dict):
    sep = "=" * 60
    print(f"\n{sep}")
    print("PREPROCESSING & K-MEANS CLUSTERING SUMMARY")
    print(sep)
    print(f"Total Sheet Diproses  : {stats.get('sheets_diproses', '-')}")
    print(f"Total Data Valid      : {stats.get('total_valid', '-')}")
    print(f"Baris Dihapus         : {stats.get('baris_dihapus', '-')}")
    print()
    print(f"Umur Invalid          : {stats.get('umur_invalid', '-')}")
    print(f"IMT Invalid           : {stats.get('imt_invalid', '-')}")
    print(f"TD Invalid            : {stats.get('td_invalid', '-')}")
    print()
    print(f"Jumlah Cluster        : {N_CLUSTERS}")
    print(f"Silhouette Score      : {metrics.get('silhouette_score', '-')}")
    print(f"Davies-Bouldin Index  : {metrics.get('davies_bouldin_index', '-')}")
    print(f"Inertia               : {metrics.get('inertia', '-')}")
    print()
    dist = metrics.get("jumlah_data_per_cluster", {})
    for cl, cnt in sorted(dist.items()):
        print(f"  Cluster {cl}          : {cnt} lansia")
    print()
    print("Output:")
    print("  - outputs/hasil_preprocessing.xlsx")
    print("  - outputs/hasil_preprocessing_scaled.xlsx")
    print("  - outputs/hasil_clustering.xlsx")
    print("  - csv/hasil_cluster_final.csv")
    print("  - outputs/elbow_method.png")
    print("  - outputs/cluster_visualization.png")
    print("  - visualisasi/  (6 grafik tambahan)")
    print("  - outputs/evaluasi_clustering.txt")
    print("  - outputs/validation_summary.txt")
    print(sep)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    global logger

    # Buat direktori output
    for d in [OUTPUT_DIR, VIS_DIR, LOG_DIR, CSV_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    logger = setup_logger("lansia", LOG_DIR)
    logger.info("=" * 60)
    logger.info("MULAI: Project Clustering Lansia Puskesmas Maesan")
    logger.info("=" * 60)

    # ── Cek file input ──
    if not INPUT_FILE.exists():
        logger.error(f"File tidak ditemukan: {INPUT_FILE}")
        sys.exit(1)
    logger.info(f"File input: {INPUT_FILE}")

    # ── TAHAP 1: Preprocessing ──
    logger.info("TAHAP 1: Preprocessing data...")
    df_clean, stats = run_preprocessing(INPUT_FILE)

    if df_clean.empty:
        logger.error("Preprocessing menghasilkan data kosong. Hentikan.")
        sys.exit(1)

    save_preprocessing_excel(df_clean, OUTPUT_DIR / "hasil_preprocessing.xlsx")
    save_validation_txt(stats, OUTPUT_DIR / "validation_summary.txt")

    # ── TAHAP 2: Clustering ──
    logger.info("TAHAP 2: K-Means Clustering...")
    (df_result, df_scaled, df_centroids,
     elbow_inertias, df_k_evaluation, metrics, interpretasi,
     X_pca, pca_variance, features) = run_clustering(df_clean)

    save_clustering_excel(df_result, df_centroids,
                          OUTPUT_DIR / "hasil_clustering.xlsx")
    save_scaled_excel(df_scaled, df_result["cluster"],
                      OUTPUT_DIR / "hasil_preprocessing_scaled.xlsx")
    save_evaluation_txt(metrics, interpretasi,
                        OUTPUT_DIR / "evaluasi_clustering.txt")
    save_k_evaluation_excel(df_k_evaluation,
                        OUTPUT_DIR / "evaluasi_semua_k.xlsx")
    save_gis_csv(df_result, CSV_DIR / "hasil_cluster_final.csv")

    # ── TAHAP 3: Visualisasi ──
    logger.info("TAHAP 3: Membuat visualisasi...")
    run_all_visualizations(df_result,df_scaled.values, X_pca, elbow_inertias,df_k_evaluation, features,
                           OUTPUT_DIR, VIS_DIR)

    # ── Summary ──
    print_summary(stats, metrics)
    logger.info("SELESAI. Semua output tersedia di folder outputs/, visualisasi/, csv/")


if __name__ == "__main__":
    main()

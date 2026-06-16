"""
preprocessing.py - Pembersihan & normalisasi data Excel lansia Puskesmas
Menggunakan openpyxl untuk membaca data mentah (menangani merged cells).
Fixed column index mapping (0-based) sesuai format laporan Posyandu Lansia.
"""

# Standard Library
from collections import defaultdict
import logging
from pathlib import Path
import re

# Third Party
import numpy as np
import openpyxl
import pandas as pd

# Local
from app.core.config import SKIP_SHEETS, MIN_FILL_RATIO

logger = logging.getLogger("lansia.preprocessing")


# ══════════════════════════════════════════════════════════════════════════════
# Konstanta: mapping posisi kolom tetap berdasarkan index (0-based)
# ══════════════════════════════════════════════════════════════════════════════

COL = {
    "nama":         1,
    "nik":          2,
    "kunjungan_baru": 3,
    "kunjungan_lama": 4,
    "umur_45_59_L": 5,
    "umur_45_59_P": 6,
    "umur_60_69_L": 7,
    "umur_60_69_P": 8,
    "umur_70p_L":   9,
    "umur_70p_P":   10,
    "imt_L":        11,
    "imt_N":        12,
    "imt_K":        13,
    "td_T":         14,
    "td_N":         15,
    "td_R":         16,
    "kemandirian_A":        17,
    "kemandirian_B_ringan": 18,
    "kemandirian_B_sedang": 19,
    "kemandirian_B_berat":  20,
    "kemandirian_B_total":  21,
    "skilas_penurunan_kognitif":     22,
    "skilas_keterbatasan_mobilitas": 23,
    "skilas_malnutrisi":             24,
    "skilas_ggn_penglihatan":        25,
    "skilas_ggn_pendengaran":        26,
    "skilas_gejala_depresi":         27,
    "hb_N":         28,
    "hb_K":         29,
    "kolesterol_N": 30,
    "kolesterol_T": 31,
    "gula_darah_N": 32,
    "gula_darah_T": 33,
    "asam_urat_N":  34,
    "asam_urat_T":  35,
    "ggn_paru":     36,
    "ggn_ginjal":   37,
    "diobati":      38,
    "dirujuk":      39,
}

# Fitur numerik utama (untuk imputasi & statistik)
NUMERIC_FEATURES = ["umur", "imt", "sistol", "diastol",
                    "hb", "kolesterol", "gula_darah", "asam_urat"]

# Semua fitur yang akan digunakan untuk K-Means
ALL_FEATURE_COLS = [
    "umur", "jenis_kelamin", "imt", "sistol", "diastol",
    "kemandirian_A", "kemandirian_B",
    "skilas_penurunan_kognitif", "skilas_keterbatasan_mobilitas",
    "skilas_malnutrisi", "skilas_ggn_penglihatan",
    "skilas_ggn_pendengaran", "skilas_gejala_depresi",
    "hb", "kolesterol", "gula_darah", "asam_urat",
    "gangguan_paru", "gangguan_ginjal",
    "kunjungan_baru", "kunjungan_lama",
    "diobati", "dirujuk",
]


# ══════════════════════════════════════════════════════════════════════════════
# Helper Functions
# ══════════════════════════════════════════════════════════════════════════════

def _safe_get(row, idx):
    """Ambil nilai dari row berdasarkan index, return None jika out of range."""
    return row[idx] if idx < len(row) else None


def _clean_str(val):
    """Bersihkan string: hapus whitespace, backtick, newline."""
    if val is None:
        return None
    s = str(val).strip().replace("`", "").replace("\n", " ").strip()
    return s if s and s.upper() not in ("NONE", "NAN") else None


def _is_data_row(row):
    """
    Baris data valid jika:
    - col[0] bisa jadi angka > 0 (nomor urut)
    - col[1] harus string nama (bukan angka) — menyaring baris header nomor kolom
    """
    if not row or row[0] is None or row[1] is None:
        return False
    try:
        float(str(row[0]).strip())
    except (ValueError, TypeError):
        return False
    # col[1] harus teks nama, bukan angka
    try:
        float(str(row[1]).strip())
        return False  # col[1] adalah angka → baris header nomor kolom, SKIP
    except (ValueError, TypeError):
        pass
    # Harus ada nama
    nama = _clean_str(row[1])
    return bool(nama)


def _extract_umur_and_gender(row):
    """
    Ekstrak umur dan jenis kelamin dari kolom usia_*_L/P.
    Return (umur: float, jenis_kelamin: int) — 1=L, 0=P.
    """
    cols = [
        (COL["umur_45_59_L"], 1),
        (COL["umur_45_59_P"], 0),
        (COL["umur_60_69_L"], 1),
        (COL["umur_60_69_P"], 0),
        (COL["umur_70p_L"],   1),
        (COL["umur_70p_P"],   0),
    ]
    for col_idx, gender in cols:
        val = _safe_get(row, col_idx)
        if val is not None:
            s = str(val).strip()
            if s and s.upper() not in ("NONE", ""):
                nums = re.findall(r"[\d.]+", s.replace(",", "."))
                if nums:
                    try:
                        return float(nums[0]), gender
                    except ValueError:
                        pass
    return None, None


def _parse_imt(bb_str):
    """
    Parse IMT dari berbagai format:
    - BB/TB: "66/160" → hitung IMT = BB / (TB/100)²
    - Angka campuran: "27/G" → ambil angka pertama
    - Angka langsung: "25.3"
    """
    if bb_str is None:
        return None
    s = str(bb_str).strip().replace(",", ".")
    if not s or s.upper() in ("NONE", "NAN", ""):
        return None
    # Format BB/TB: dua angka dipisah slash
    slash = re.match(r"^([\d.]+)\s*/\s*([\d.]+)$", s)
    if slash:
        try:
            bb = float(slash.group(1))
            tb = float(slash.group(2))
            if tb > 3:
                tb /= 100
            if tb > 0 and bb > 0:
                return round(bb / (tb ** 2), 2)
        except (ValueError, ZeroDivisionError):
            pass
    # Format '27/G' → ambil angka pertama
    slash_mix = re.match(r"^([\d.]+)\s*/", s)
    if slash_mix:
        try:
            v = float(slash_mix.group(1))
            return round(v, 2) if v > 5 else None
        except ValueError:
            pass
    # Angka langsung
    nums = re.findall(r"[\d.]+", s)
    if nums:
        try:
            return round(float(nums[0]), 2)
        except ValueError:
            pass
    return None


def _parse_td(row):
    """Parse tekanan darah sistolik/diastolik dari format '120/80'."""
    for col_idx in [COL["td_T"], COL["td_N"], COL["td_R"]]:
        val = _safe_get(row, col_idx)
        if val is not None:
            s = str(val).strip().replace(",", ".")
            m = re.match(r"^(\d+)\s*/\s*(\d+)$", s)
            if m:
                return int(m.group(1)), int(m.group(2))
    return None, None


def _has_value(row, col_idx):
    """Cek apakah sel terisi (bukan kosong/0/None)."""
    val = _safe_get(row, col_idx)
    if val is None:
        return 0
    s = str(val).strip().upper()
    return 0 if s in ("", "NONE", "NAN", "0", "0.0") else 1


def _parse_skilas_val(row, col_idx):
    """
    Parse nilai SKILAS:
    - Y, G, V, 1 → 1
    - N, -, T, kosong, atau 0 → 0
    """
    val = _safe_get(row, col_idx)
    if val is None:
        return 0
    s = str(val).strip().upper()
    if s.endswith(".0"):
        s = s[:-2]
    if s in ("Y", "G", "V", "1"):
        return 1
    if s in ("N", "-", "T", "", "NONE", "NAN", "0"):
        return 0
    # Fallback jika berupa angka positif
    try:
        if float(s) > 0:
            return 1
    except ValueError:
        pass
    return 0



def _get_numeric(row, *col_indices):
    """Ambil nilai numerik pertama yang valid dari beberapa kolom."""
    for col_idx in col_indices:
        val = _safe_get(row, col_idx)
        if val is not None:
            try:
                v = float(str(val).strip().replace(",", "."))
                if v > 0:
                    return v
            except (ValueError, TypeError):
                pass
    return None


def _parse_binary_field(row, col_idx):
    """Parse field biner: terisi → 1, kosong → 0."""
    val = _safe_get(row, col_idx)
    if val is None:
        return 0
    s = str(val).strip().upper()
    return 0 if s in ("", "NONE", "NAN", "0", "0.0", "T") else 1


# ══════════════════════════════════════════════════════════════════════════════
# Sheet Processing
# ══════════════════════════════════════════════════════════════════════════════

def _process_sheet(ws, sheet_name):
    """Proses satu sheet worksheet openpyxl menjadi list of dict."""
    records = []
    for row in ws.iter_rows(values_only=True):
        if not _is_data_row(row):
            continue

        row = list(row) + [None] * 60  # pad agar aman akses index

        nama = _clean_str(row[COL["nama"]])
        if not nama:
            continue

        nik = _clean_str(row[COL["nik"]])

        # Kunjungan: pisahkan Baru dan Lama
        baru_val = _safe_get(row, COL["kunjungan_baru"])
        lama_val = _safe_get(row, COL["kunjungan_lama"])

        def is_filled(v):
            if v is None:
                return False
            s = str(v).strip().upper()
            return s not in ("", "NONE", "NAN", "0", "0.0")

        if is_filled(baru_val):
            kunjungan_baru = 1
            kunjungan_lama = 0
        elif is_filled(lama_val):
            kunjungan_baru = 0
            kunjungan_lama = 1
        else:
            kunjungan_baru = 0
            kunjungan_lama = 0

        umur, jenis_kelamin = _extract_umur_and_gender(row)

        imt_raw = (_safe_get(row, COL["imt_L"])
                   or _safe_get(row, COL["imt_N"])
                   or _safe_get(row, COL["imt_K"]))
        imt = _parse_imt(imt_raw)

        sistol, diastol = _parse_td(row)

        kemandirian_A = _has_value(row, COL["kemandirian_A"])
        kemandirian_B = max(
            _has_value(row, COL["kemandirian_B_ringan"]),
            _has_value(row, COL["kemandirian_B_sedang"]),
            _has_value(row, COL["kemandirian_B_berat"]),
            _has_value(row, COL["kemandirian_B_total"]),
        )

        hb         = _get_numeric(row, COL["hb_N"], COL["hb_K"])
        kolesterol  = _get_numeric(row, COL["kolesterol_N"], COL["kolesterol_T"])
        gula_darah  = _get_numeric(row, COL["gula_darah_N"], COL["gula_darah_T"])
        asam_urat   = _get_numeric(row, COL["asam_urat_N"], COL["asam_urat_T"])

        records.append({
            "desa":               sheet_name,
            "nama":               nama,
            "nik":                nik,
            "kunjungan_baru":     kunjungan_baru,
            "kunjungan_lama":     kunjungan_lama,
            "umur":               umur,
            "jenis_kelamin":      jenis_kelamin,
            "imt":                imt,
            "sistol":             sistol,
            "diastol":            diastol,
            "kemandirian_A":      kemandirian_A,
            "kemandirian_B":      kemandirian_B,
            "skilas_penurunan_kognitif":     _parse_skilas_val(row, COL["skilas_penurunan_kognitif"]),
            "skilas_keterbatasan_mobilitas": _parse_skilas_val(row, COL["skilas_keterbatasan_mobilitas"]),
            "skilas_malnutrisi":             _parse_skilas_val(row, COL["skilas_malnutrisi"]),
            "skilas_ggn_penglihatan":        _parse_skilas_val(row, COL["skilas_ggn_penglihatan"]),
            "skilas_ggn_pendengaran":        _parse_skilas_val(row, COL["skilas_ggn_pendengaran"]),
            "skilas_gejala_depresi":         _parse_skilas_val(row, COL["skilas_gejala_depresi"]),
            "hb":                 hb,
            "kolesterol":         kolesterol,
            "gula_darah":         gula_darah,
            "asam_urat":          asam_urat,
            "gangguan_paru":      _parse_binary_field(row, COL["ggn_paru"]),
            "gangguan_ginjal":    _parse_binary_field(row, COL["ggn_ginjal"]),
            "diobati":            _parse_binary_field(row, COL["diobati"]),
            "dirujuk":            _parse_binary_field(row, COL["dirujuk"]),
        })

    return records


def _load_all_sheets(filepath):
    """Baca semua sheet desa dari workbook Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_records = []
    for name in wb.sheetnames:
        if name.strip() in SKIP_SHEETS or name.lower().strip() in {s.lower() for s in SKIP_SHEETS}:
            logger.info(f"  ⏭  Skip: {name}")
            continue
        recs = _process_sheet(wb[name], name)
        logger.info(f"  ✅ {name:<20}: {len(recs):>4} baris")
        all_records.extend(recs)
    return pd.DataFrame(all_records)


# ══════════════════════════════════════════════════════════════════════════════
# Cleaning & Validasi Range
# ══════════════════════════════════════════════════════════════════════════════

def _clean_df(df):
    """Validasi range dan normalisasi tipe data."""
    df = df[df["nama"].notna()].copy()

    # Validasi range: nilai di luar batas → NaN
    ranges = {
        "umur":       (40, 110),
        "imt":        (10, 60),
        "sistol":     (60, 300),
        "diastol":    (40, 200),
        "hb":         (5, 25),
        "kolesterol": (50, 600),
        "gula_darah": (30, 800),
        "asam_urat":  (1, 25),
    }
    for col, (lo, hi) in ranges.items():
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df.loc[df[col] < lo, col] = np.nan
            df.loc[df[col] > hi, col] = np.nan

    # Jenis kelamin: NaN → 0 (default perempuan)
    df["jenis_kelamin"] = df["jenis_kelamin"].fillna(0).astype(int)

    # Kolom biner: NaN → 0
    binary = [
        "kunjungan_baru", "kunjungan_lama", "kemandirian_A", "kemandirian_B",
        "skilas_penurunan_kognitif", "skilas_keterbatasan_mobilitas",
        "skilas_malnutrisi", "skilas_ggn_penglihatan",
        "skilas_ggn_pendengaran", "skilas_gejala_depresi",
        "gangguan_paru", "gangguan_ginjal", "diobati", "dirujuk",
    ]
    for col in binary:
        df[col] = df[col].fillna(0).astype(int)

    return df.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# Drop Kolom Seluruhnya Null/0
# ══════════════════════════════════════════════════════════════════════════════

def _drop_all_null_columns(df, label=""):
    """Hapus kolom yang seluruh nilainya null/0 di semua desa."""
    numeric_cols = ["hb", "kolesterol", "gula_darah", "asam_urat",
                    "sistol", "diastol", "imt", "umur"]
    dropped = []
    for col in numeric_cols:
        if col in df.columns and df[col].isna().all():
            dropped.append(col)

    # Kolom biner: drop jika semua 0
    binary_cols = [
        "skilas_penurunan_kognitif", "skilas_keterbatasan_mobilitas",
        "skilas_malnutrisi", "skilas_ggn_penglihatan",
        "skilas_ggn_pendengaran", "skilas_gejala_depresi",
        "gangguan_paru", "gangguan_ginjal", "kemandirian_A", "kemandirian_B",
        "diobati", "dirujuk",
    ]
    for col in binary_cols:
        if col in df.columns and df[col].sum() == 0:
            dropped.append(col)

    if dropped:
        logger.info(f"🗑  Kolom dihapus (semua null/nol di {label}): {dropped}")
        df = df.drop(columns=dropped)
    else:
        logger.info(f"✅ Tidak ada kolom yang sepenuhnya null di {label}")
    return df, dropped


# ══════════════════════════════════════════════════════════════════════════════
# Imputasi Missing Values
# ══════════════════════════════════════════════════════════════════════════════

def _impute_for_kmeans(df, feature_cols):
    """
    Imputasi kolom numerik per desa dengan aturan threshold:
    - Jika kolom terisi >= MIN_FILL_RATIO di suatu desa → imputasi missing dengan median desa
    - Jika kolom terisi < MIN_FILL_RATIO di suatu desa  → SKIP, biarkan NaN (data terlalu sedikit)
    
    Tidak ada global fallback — desa tanpa data cukup tetap kosong untuk kolom itu.
    """
    
    num_in_features = [c for c in NUMERIC_FEATURES if c in feature_cols]
    df_imputed = df.copy()
    
    for col in num_in_features:
        skipped_desa = []
        imputed_desa = []
        
        for desa, group in df.groupby("desa"):
            non_null = group[col].notna().sum()
            total    = len(group)
            ratio    = non_null / total if total > 0 else 0
            
            if ratio >= MIN_FILL_RATIO:
                # Cukup data → imputasi NaN dengan median desa
                med = group[col].median()
                if pd.notna(med):
                    n_missing = group[col].isna().sum()
                    if n_missing > 0:
                        idx = df_imputed[
                            (df_imputed["desa"] == desa) & (df_imputed[col].isna())
                        ].index
                        df_imputed.loc[idx, col] = med
                        imputed_desa.append(f"{desa}({non_null}/{total}, isi {n_missing})")
            else:
                # Terlalu sedikit data → skip, biarkan NaN
                if non_null > 0:
                    skipped_desa.append(f"{desa}({non_null}/{total})")
                else:
                    skipped_desa.append(f"{desa}(kosong)")
        
        if skipped_desa:
            logger.info(f"  {col:<15}: SKIP desa {', '.join(skipped_desa)}")
        if imputed_desa:
            logger.info(f"  {col:<15}: IMPUTASI desa {', '.join(imputed_desa)}")
                    
    return df_imputed


# ══════════════════════════════════════════════════════════════════════════════
# Pipeline utama preprocessing
# ══════════════════════════════════════════════════════════════════════════════

def run_preprocessing(input_file) -> tuple:
    """
    Jalankan seluruh pipeline preprocessing.

    Kembalikan:
        (df_clean, stats_dict)
        df_clean : DataFrame gabungan semua desa, sudah di-clean dan diimputasi
        stats    : dict statistik untuk laporan akhir
    """
    input_file = Path(input_file) if not isinstance(input_file, Path) else input_file

    logger.info("=" * 60)
    logger.info("  PREPROCESSING DATA SKILAS → K-MEANS")
    logger.info("=" * 60)

    stats = defaultdict(int)

    # ── Tahap 1: Baca semua sheet ──
    logger.info("📂 Membaca data dari semua sheet desa...")
    df = _load_all_sheets(input_file)
    stats["total_baris_mentah"] = len(df)
    logger.info(f"📊 Total baris terbaca: {len(df)}")

    if df.empty:
        logger.error("Tidak ada data yang berhasil diproses!")
        return pd.DataFrame(), dict(stats)

    # Hitung jumlah sheet
    stats["sheets_diproses"] = int(df["desa"].nunique())

    # ── Tahap 2: Cleaning & validasi range ──
    logger.info("🧹 Cleaning & validasi range...")
    df = _clean_df(df)
    stats["total_valid"] = len(df)
    stats["baris_dihapus"] = stats["total_baris_mentah"] - stats["total_valid"]
    logger.info(f"📊 Total setelah cleaning: {len(df)}")

    # Distribusi kunjungan
    logger.info(f"🔍 Distribusi kunjungan: Baru={df['kunjungan_baru'].sum()}, "
                f"Lama={df['kunjungan_lama'].sum()}")

    # Missing values info
    logger.info("📉 Missing values kolom numerik (setelah cleaning):")
    for col in NUMERIC_FEATURES:
        if col in df.columns:
            n = df[col].isna().sum()
            pct = n / len(df) * 100 if len(df) > 0 else 0
            logger.info(f"  {col:<15}: {n:>4} missing ({pct:.1f}%)")

    # Statistik missing sebelum imputasi
    stats["umur_invalid"] = int(df["umur"].isna().sum()) if "umur" in df.columns else 0
    stats["imt_invalid"] = int(df["imt"].isna().sum()) if "imt" in df.columns else 0
    stats["td_invalid"] = int(df["sistol"].isna().sum()) if "sistol" in df.columns else 0

    # ── Tahap 3: Imputasi missing values ──
    feature_cols = ALL_FEATURE_COLS
    logger.info("🔢 Imputasi missing values (median per desa)...")
    df = _impute_for_kmeans(df, feature_cols)

    # ── Tahap 4: Mengisi sisa data kosong (NaN) dengan 0 ──
    logger.info("🧹 Mengisi sisa data kosong (NaN) dengan 0...")
    df[feature_cols] = df[feature_cols].fillna(0)

    still_nan = df[feature_cols].isna().sum().sum()
    logger.info(f"   Sisa NaN di fitur K-Means: {still_nan}")

    # Statistik akhir
    logger.info(f"📋 Jumlah data per desa:")
    for desa, count in df.groupby("desa").size().items():
        logger.info(f"  {desa}: {count}")

    logger.info(f"📌 Fitur K-Means yang digunakan ({len(feature_cols)} kolom): "
                f"{', '.join(feature_cols)}")

    logger.info(f"Total data gabungan: {len(df)} baris, {len(df.columns)} kolom")
    return df, dict(stats)
